"""
generar_informe.py
------------------
Genera un informe Excel de participación en encuesta.

Para ejecutar:
    python generar_informe.py

Configuración: edita el archivo .env en el mismo directorio.
"""

import sys
import re
import os
from pathlib import Path

from dotenv import load_dotenv
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# =============================================================================
#  Carga de configuración desde .env
# =============================================================================

load_dotenv()

def _cargar_config():
    empresa   = os.getenv("EMPRESA", "").strip()
    salida    = os.getenv("SALIDA", "informe.xlsx").strip()
    separador = os.getenv("SEPARADOR", ";").strip()

    publicos_y_csvs = []
    i = 1
    while True:
        publico = os.getenv(f"PUBLICO_{i}", "").strip()
        csv     = os.getenv(f"CSV_{i}", "").strip()
        if not publico and not csv:
            break
        if not publico or not csv:
            sys.exit(f"[ERROR] .env: PUBLICO_{i} y CSV_{i} deben definirse juntos.")
        publicos_y_csvs.append((publico, csv))
        i += 1

    if not empresa:
        sys.exit("[ERROR] .env: EMPRESA es obligatorio.")
    if not publicos_y_csvs:
        sys.exit("[ERROR] .env: define al menos PUBLICO_1 y CSV_1.")

    return empresa, salida, separador, publicos_y_csvs

EMPRESA, SALIDA, SEPARADOR, PUBLICOS_Y_CSVS = _cargar_config()

# =============================================================================


# ── estilos ───────────────────────────────────────────────────────────────────

THIN   = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADER_FILL  = PatternFill("solid", start_color="4472C4")
PUBLICO_FILL = PatternFill("solid", start_color="D9E1F2")
HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
PUBLICO_FONT = Font(bold=True, name="Arial", size=10)
CELL_FONT    = Font(name="Arial", size=10)
TITLE_FONT   = Font(bold=True, name="Arial", size=11)
CENTER       = Alignment(horizontal="center", vertical="center")
LEFT         = Alignment(horizontal="left",   vertical="center")


def style_cell(cell, value=None, font=None, fill=None, alignment=None,
                border=None, number_format=None):
    if value is not None:
        cell.value = value
    if font:          cell.font          = font
    if fill:          cell.fill          = fill
    if alignment:     cell.alignment     = alignment
    if border:        cell.border        = border
    if number_format: cell.number_format = number_format


# ── extracción de fecha desde el nombre del archivo ───────────────────────────

def fecha_desde_nombre(csv_path: str) -> str:
    """
    Extrae la fecha del nombre del archivo.
    Formato esperado: ..._YYYYMMDDHHMMSS (14 dígitos tras el último '_').
    Devuelve la fecha como 'dd/mm'.
    """
    stem = Path(csv_path).stem
    m = re.search(r"(\d{8})\d{6}", stem)
    if not m:
        m = re.search(r"(\d{8})", stem)
    if not m:
        sys.exit(
            f"[ERROR] No se pudo extraer la fecha del nombre '{csv_path}'.\n"
            f"        El nombre debe contener la fecha en formato YYYYMMDD (p.ej. 20260507)."
        )
    raw = m.group(1)   # 'YYYYMMDD'
    return f"{raw[6:8]}/{raw[4:6]}"


# ── carga y procesado del CSV ─────────────────────────────────────────────────

def cargar_csv(path: str) -> pd.DataFrame:
    if not Path(path).exists():
        sys.exit(f"[ERROR] Archivo no encontrado: \'{path}\'")

    required = {"IdUsuario", "Nombre", "Empresa", "Estado"}

    # Estrategias de lectura en orden de preferencia
    estrategias = [
        dict(sep=SEPARADOR, quotechar='"',  engine="python", dtype=str),
        dict(sep=SEPARADOR, quotechar="'", engine="python", dtype=str),
        dict(sep=SEPARADOR, quoting=3,      engine="python", dtype=str),  # QUOTE_NONE
        dict(sep=SEPARADOR,                 engine="python", dtype=str),
        dict(sep=SEPARADOR, quotechar='"',  engine="c", dtype=str, on_bad_lines="skip"),
        dict(sep=SEPARADOR, quoting=3,      engine="c", dtype=str, on_bad_lines="skip"),
    ]

    last_error = None
    for kwargs in estrategias:
        try:
            df = pd.read_csv(path, **kwargs).fillna("")
            # Limpiar comillas residuales de cabeceras (pasa con QUOTE_NONE)
            df.columns = [c.strip().strip('"').strip("'") for c in df.columns]
            # Limpiar comillas residuales de valores de texto
            for col in df.select_dtypes(include="string").columns:
                df[col] = df[col].str.strip().str.strip('"').str.strip("'")
            if required.issubset(set(df.columns)):
                return df
        except Exception as e:
            last_error = e
            continue

    # Ninguna estrategia funcionó — mostrar diagnóstico
    try:
        with open(path, "rb") as f:
            head = f.read(512)
        print(f"[DIAGNÓSTICO] Primeros bytes:\n{head[:200]}")
    except Exception:
        pass
    sys.exit(
        f"[ERROR] No se pudo leer \'{path}\' con ninguna estrategia de parseo.\n"
        f"        Último error: {last_error}\n"
        f"        Comprueba que SEPARADOR=\'{SEPARADOR}\' es correcto."
    )


def empresa_desde_csv(df: pd.DataFrame, path: str) -> str:
    """Toma el valor de la columna Empresa de la primera fila no vacía."""
    no_vacios = df["Empresa"].str.strip()
    no_vacios = no_vacios[no_vacios != ""]
    if no_vacios.empty:
        sys.exit(f"[ERROR] La columna 'Empresa' está vacía en '{path}'.")
    return no_vacios.iloc[0]


def procesar_publico(df: pd.DataFrame) -> pd.DataFrame:
    """
    Devuelve DataFrame con ['Nombre', 'Estado']:
        - Filtra por EMPRESA
        - Excluye filas con Estado == 'Anulado'
        - Si Nombre == 'Lote' → usa IdUsuario como Nombre
        - Si no              → usa Nombre tal cual
    """
    df = df[df["Empresa"].str.strip() == EMPRESA.strip()]
    df = df[df["Estado"].str.strip() != "Anulado"]
    rows = []
    for _, r in df.iterrows():
        nombre = r["IdUsuario"] if r["Nombre"].strip() == "Lote" else r["Nombre"]
        rows.append({"Nombre": nombre, "Estado": r["Estado"]})
    return pd.DataFrame(rows)


# ── construcción del Excel ────────────────────────────────────────────────────

def build_excel(empresa: str, fecha: str, publicos: list,
                dataframes: list, salida: str) -> None:

    wb = Workbook()
    ws = wb.active
    ws.title = "Informe"

    ws.row_dimensions[1].height = 6

    # B2: título
    titulo = f"A FECHA {fecha} {empresa}"
    c = ws.cell(row=2, column=2, value=titulo)
    c.font      = TITLE_FONT
    c.alignment = LEFT
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 18

    # Layout: 3 columnas por público + 1 separadora entre ellos
    COL_START  = 3
    COLS_X_PUB = 3
    SEPARATOR  = 0

    col_offsets = [COL_START + i * (COLS_X_PUB + SEPARATOR) for i in range(len(publicos))]

    # Filas 2-3: nombre del público y cabeceras
    for pub_name, col_start in zip(publicos, col_offsets):
        c_nombre, c_estado, c_pct = col_start, col_start + 1, col_start + 2

        ws.merge_cells(start_row=2, start_column=c_nombre,
                        end_row=2,   end_column=c_pct)
        style_cell(ws.cell(row=2, column=c_nombre), pub_name,
                    font=PUBLICO_FONT, fill=PUBLICO_FILL,
                    alignment=CENTER, border=BORDER)

        for col, header in [(c_nombre, "Nombre"),
                            (c_estado, "Estado"),
                            (c_pct,    "% Participación")]:
            style_cell(ws.cell(row=3, column=col), header,
                        font=HEADER_FONT, fill=HEADER_FILL,
                        alignment=CENTER, border=BORDER)

    # Filas de datos (fila 4 en adelante)
    DATA_START = 4

    for pub_df, col_start in zip(dataframes, col_offsets):
        c_nombre = col_start
        c_estado = col_start + 1
        c_pct    = col_start + 2
        data_end     = DATA_START + len(pub_df) - 1
        letra_estado = get_column_letter(c_estado)

        for row_idx, (_, row) in enumerate(pub_df.iterrows()):
            excel_row = DATA_START + row_idx

            style_cell(ws.cell(row=excel_row, column=c_nombre), row["Nombre"],
                        font=CELL_FONT, alignment=LEFT, border=BORDER)

            style_cell(ws.cell(row=excel_row, column=c_estado), row["Estado"],
                        font=CELL_FONT, alignment=CENTER, border=BORDER)

            formula = (
                f"=COUNTIF(${letra_estado}${DATA_START}:${letra_estado}${data_end},"
                f"{letra_estado}{excel_row})"
                f"/COUNTA(${letra_estado}${DATA_START}:${letra_estado}${data_end})"
            )
            style_cell(ws.cell(row=excel_row, column=c_pct), formula,
                        font=CELL_FONT, alignment=CENTER, border=BORDER,
                        number_format="0%")

    # Anchos de columna
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 32

    for i, col_start in enumerate(col_offsets):
        ws.column_dimensions[get_column_letter(col_start)].width     = 22
        ws.column_dimensions[get_column_letter(col_start + 1)].width = 12
        ws.column_dimensions[get_column_letter(col_start + 2)].width = 16
        if i < len(col_offsets) - 1:
            ws.column_dimensions[get_column_letter(col_start + 3)].width = 2

    try:
        wb.save(salida)
        print(f"[OK] Informe guardado en: {salida}")
    except Exception as e:
        sys.exit(f"[ERROR] No se pudo guardar '{salida}': {e}")


# ── main ──────────────────────────────────────────────────────────────────────

def main():
    if not PUBLICOS_Y_CSVS:
        sys.exit("[ERROR] PUBLICOS_Y_CSVS está vacío. Edita la configuración.")
    if not EMPRESA.strip():
        sys.exit("[ERROR] EMPRESA está vacío. Edita la configuración.")

    publicos   = []
    dataframes = []
    fecha      = None

    for pub_nombre, csv_path in PUBLICOS_Y_CSVS:
        raw = cargar_csv(csv_path)

        if fecha is None:
            fecha = fecha_desde_nombre(csv_path)

        df = procesar_publico(raw)
        if df.empty:
            print(f"[AVISO] '{pub_nombre}': ningún contacto encontrado para '{EMPRESA}'. Se omite.")
            continue

        publicos.append(pub_nombre)
        dataframes.append(df)

    if not publicos:
        sys.exit(f"[ERROR] No se encontraron contactos para '{EMPRESA}' en ningún CSV.")

    print(f"[INFO] Empresa : {EMPRESA}")
    print(f"[INFO] Fecha   : {fecha}")
    print(f"[INFO] Públicos: {', '.join(publicos)}")

    build_excel(EMPRESA, fecha, publicos, dataframes, SALIDA)


if __name__ == "__main__":
    main()