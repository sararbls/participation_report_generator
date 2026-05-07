import pandas as pd
from openpyxl import load_workbook

from participation_report.excel_builder import build_excel


def test_build_excel_crea_hoja_titulo_headers_y_formulas(tmp_path):
    salida = tmp_path / "informe.xlsx"
    publicos = ["Clientes"]
    dataframes = [
        pd.DataFrame(
            [
                {"Nombre": "Ana", "Estado": "Completa"},
                {"Nombre": "u2", "Estado": "Pendiente"},
            ]
        )
    ]

    build_excel("Hospital Central", "07/05", publicos, dataframes, str(salida))

    assert salida.exists()

    wb = load_workbook(str(salida))
    ws = wb["Informe"]

    assert ws["B2"].value == "A FECHA 07/05 Hospital Central"
    assert ws["C2"].value == "Clientes"
    assert ws["C3"].value == "Nombre"
    assert ws["D3"].value == "Estado"
    assert ws["E3"].value == "% Participación"

    assert ws["C4"].value == "Ana"
    assert ws["D4"].value == "Completa"
    assert ws["E4"].value == "=COUNTIF($D$4:$D$5,D4)/COUNTA($D$4:$D$5)"
    assert ws["E4"].number_format == "0%"
