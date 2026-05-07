import re

from openpyxl import load_workbook

from participation_report.config import AppConfig
from participation_report.services import generate_report


def test_generate_report_end_to_end_crea_excel_en_processed_data(
    tmp_path, monkeypatch, write_contacts_csv, capsys
):
    monkeypatch.chdir(tmp_path)

    csv_path = write_contacts_csv(
        "contactos_20260507101650.csv",
        [
            {
                "IdUsuario": "u1",
                "Nombre": "Ana",
                "Empresa": "Hospital Central",
                "Estado": "Completa",
            },
            {
                "IdUsuario": "u2",
                "Nombre": "Lote",
                "Empresa": "Hospital Central",
                "Estado": "Pendiente",
            },
            {
                "IdUsuario": "u3",
                "Nombre": "Luis",
                "Empresa": "Hospital Central",
                "Estado": "Anulado",
            },
        ],
    )

    cfg = AppConfig(
        empresa="Hospital Central",
        separador=";",
        publicos_y_csvs=[("Clientes", str(csv_path))],
    )

    generate_report(cfg)
    out = capsys.readouterr().out

    assert "[INFO] Empresa : Hospital Central" in out
    assert "[INFO] Fecha   : 07/05" in out
    ok_pattern = (
        r"\[OK\] Informe guardado en: processed_data[\\/]"
        r"reporte_participacion_Hospital_Central_\d{8}_\d{6}\.xlsx"
    )
    assert re.search(ok_pattern, out)

    generated = list((tmp_path / "processed_data").glob("reporte_participacion_*.xlsx"))
    assert len(generated) == 1

    wb = load_workbook(generated[0])
    ws = wb["Informe"]
    assert ws["B2"].value == "A FECHA 07/05 Hospital Central"
    assert ws["C4"].value == "Ana"
    assert ws["C5"].value is None
