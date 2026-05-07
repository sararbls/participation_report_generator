from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADER_FILL = PatternFill("solid", start_color="4472C4")
PUBLICO_FILL = PatternFill("solid", start_color="D9E1F2")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
PUBLICO_FONT = Font(bold=True, name="Arial", size=10)
CELL_FONT = Font(name="Arial", size=10)
TITLE_FONT = Font(bold=True, name="Arial", size=11)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def style_cell(
    cell: Any,
    value: Any = None,
    font: Font | None = None,
    fill: PatternFill | None = None,
    alignment: Alignment | None = None,
    border: Border | None = None,
    number_format: str | None = None,
) -> None:
    if value is not None:
        cell.value = value
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format


def build_excel(
    empresa: str,
    fecha: str,
    publicos: list[str],
    dataframes: list[pd.DataFrame],
    salida: str,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Informe"

    ws.row_dimensions[1].height = 6

    titulo = f"A FECHA {fecha} {empresa}"
    c = ws.cell(row=2, column=2, value=titulo)
    c.font = TITLE_FONT
    c.alignment = LEFT
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 18

    col_start = 3
    cols_x_pub = 3
    separator = 0
    col_offsets = [col_start + i * (cols_x_pub + separator) for i in range(len(publicos))]

    for pub_name, start_col in zip(publicos, col_offsets):
        c_nombre, c_estado, c_pct = start_col, start_col + 1, start_col + 2

        ws.merge_cells(start_row=2, start_column=c_nombre, end_row=2, end_column=c_pct)
        style_cell(
            ws.cell(row=2, column=c_nombre),
            pub_name,
            font=PUBLICO_FONT,
            fill=PUBLICO_FILL,
            alignment=CENTER,
            border=BORDER,
        )

        for col, header in [(c_nombre, "Nombre"), (c_estado, "Estado"), (c_pct, "% Participación")]:
            style_cell(
                ws.cell(row=3, column=col),
                header,
                font=HEADER_FONT,
                fill=HEADER_FILL,
                alignment=CENTER,
                border=BORDER,
            )

    data_start = 4

    for pub_df, start_col in zip(dataframes, col_offsets):
        c_nombre = start_col
        c_estado = start_col + 1
        c_pct = start_col + 2
        data_end = data_start + len(pub_df) - 1
        letra_estado = get_column_letter(c_estado)

        for row_idx, (_, row) in enumerate(pub_df.iterrows()):
            excel_row = data_start + row_idx

            style_cell(
                ws.cell(row=excel_row, column=c_nombre),
                row["Nombre"],
                font=CELL_FONT,
                alignment=LEFT,
                border=BORDER,
            )
            style_cell(
                ws.cell(row=excel_row, column=c_estado),
                row["Estado"],
                font=CELL_FONT,
                alignment=CENTER,
                border=BORDER,
            )

            formula = (
                f"=COUNTIF(${letra_estado}${data_start}:${letra_estado}${data_end},{letra_estado}{excel_row})"
                f"/COUNTA(${letra_estado}${data_start}:${letra_estado}${data_end})"
            )
            style_cell(
                ws.cell(row=excel_row, column=c_pct),
                formula,
                font=CELL_FONT,
                alignment=CENTER,
                border=BORDER,
                number_format="0%",
            )

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 32

    for i, start_col in enumerate(col_offsets):
        ws.column_dimensions[get_column_letter(start_col)].width = 22
        ws.column_dimensions[get_column_letter(start_col + 1)].width = 12
        ws.column_dimensions[get_column_letter(start_col + 2)].width = 16
        if i < len(col_offsets) - 1:
            ws.column_dimensions[get_column_letter(start_col + 3)].width = 2

    wb.save(salida)
